import os
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from datetime import datetime
from tkinter import Tk
from tkinter.filedialog import asksaveasfilename


def gerar_excel(dados, nome_relatorio):
    """
    ! Gerar um arquivo .xlsx (Excel) a partir de uma query SQL
    """

    if not dados:
        raise ValueError("A lista de dados está vazia! nada para exportar")

    # * Cria o arquivo e define a planilha
    wb = Workbook()
    ws = wb.active
    ws.title = "Relatorio"

    # * Pega o nome das colunas automaticamente
    colunas = list(dados[0].keys())

    # !=== Cabeçalho ===
    ws.append(colunas)

    # * Estilo do cabeçalho
    header_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    header_font = Font(color="000000", bold=True)
    header_alignment = Alignment(horizontal="center", vertical="center")

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment

    # !=== LINHA DE DADOS ===
    colunas_contabeis = ["preço tabela", "custo contábl", "bonus", "custo hist.", "custo pres."]

    for linha in dados:
        ws.append(list(linha.values()))
    
    # * Formatação Contábil
    for row in ws.iter_rows(min_row=2):
        for cell, nome_coluna in zip(row, colunas):
            if nome_coluna.lower() in colunas_contabeis:
                cell.number_format = '_-* "R$" #,##0.00_-;_-* "R$" #,##0.00_-;_-* "-"??_-;_-@_-'  # ! Formato contábil Brasileiro

    # * Bordas
    thin = Side(border_style="thin", color="000000")
    for row in ws.iter_rows():
        for cell in row:
            cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)

    # !=== Ajuste automatico de largura das colunas ===
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width

    # !=== Pergunta onde salvar o arquivo ===
    root = Tk()
    root.withdraw() # Oculta a janela principal do tkinter

    timestamp = datetime.now().strftime("%d-%m-%Y")
    nome_sugerido = f"{nome_relatorio} {timestamp}.xlsx"

    caminho_salvar = asksaveasfilename(
        title="Salvar Relatório Excel",
        defaultextension=".xlsx",
        initialfile=nome_sugerido,
        filetypes=[("Arquivos Excel", "*.xlsx")],
        initialdir=os.path.expanduser("~/Documents")    # * Faz abrir direto na pasta de documentos
    )

    root.destroy()

    # * Se o usuário cancelar
    if not caminho_salvar:
        print("Operação cancelada pelo usuário")
        return
    
    # * Salva o arquivo
    wb.save(caminho_salvar)
    print(f"Arquivo salvo com sucesso em: \n{caminho_salvar}")